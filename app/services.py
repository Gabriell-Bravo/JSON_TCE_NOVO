from __future__ import annotations

import csv
import hashlib
import io
import json
from decimal import Decimal
from typing import Any

from jsonschema import Draft7Validator, FormatChecker
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import delete, func, select
from sqlalchemy.orm import Session

from .config import EXPORT_DIR, MAX_UPLOAD_SIZE
from .models import Beneficiario, Folha, FolhaItem, Programa, ProgramaCriterio, Remessa, User, ValidationIssue
from .schema_audfoben import AUDFOBEN_SCHEMA
from .catalogos import CRITERIOS_ELEGIBILIDADE, criterio_por_id
from .validators import (
    clean_text,
    cpf_valido,
    digits_only,
    gerar_criterio_unico,
    nacionalidade_deve_ser_pais,
    normalize_date,
    normalize_money,
    parse_json_array,
    validar_beneficiario_basico,
    validar_criterios,
    validar_criterios_do_programa,
    validar_dependentes,
)

EXPECTED_COLUMNS = [
    "item_id",
    "cpf", "numeroNIS", "nome", "sexo", "dataNascimento", "nacionalidade", "nomeMae", "enderecoCEP",
    "logradouro", "bairro", "numero", "complemento", "codigoIBGEMunicipio", "valorTotalTransferido",
    "totalPessoasBeneficio", "totalDependentesBeneficio", "criterio_id", "criterio_valor", "criterio_aplicavel",
    "dependentes_json", "evidencia",
]

REQUIRED_IMPORT_COLUMNS = [
    "cpf", "nome", "sexo", "dataNascimento", "nacionalidade", "logradouro", "bairro",
    "codigoIBGEMunicipio", "valorTotalTransferido", "totalPessoasBeneficio", "totalDependentesBeneficio",
]

SCHEMA_ERROR_HINTS = {
    "is a required property": "campo obrigatório ausente",
    "is not of type": "tipo de dado incorreto",
    "is not valid under any of the given schemas": "valor fora do padrão permitido",
    "is less than the minimum": "valor abaixo do mínimo permitido",
    "is greater than the maximum": "valor acima do máximo permitido",
    "does not match": "formato inválido",
}


def friendly_schema_message(error_message: str) -> str:
    for key, hint in SCHEMA_ERROR_HINTS.items():
        if key in error_message:
            return hint
    return error_message


def _criterio_dict_from_row(c: ProgramaCriterio) -> dict[str, Any]:
    return {
        "identificadorCriterio": c.identificador_criterio,
        "nome": c.nome,
        "categoria": c.categoria,
        "tipoDado": c.tipo_dado,
        "limiteInferior": c.limite_inferior or "",
        "limiteSuperior": c.limite_superior or "",
        "vigenciaInicio": c.vigencia_inicio or "",
        "vigenciaFim": c.vigencia_fim or "",
        "prazoIndeterminado": c.prazo_indeterminado,
    }


def _criterios_from_json(programa: Programa) -> list[dict[str, Any]]:
    try:
        data = json.loads(programa.criterios_padrao_json or "[]")
    except Exception:
        data = []
    if isinstance(data, list) and data:
        return [c for c in data if isinstance(c, dict)]
    return []


def _programa_criterios(programa: Programa) -> list[dict[str, Any]]:
    if programa.criterios_vinculados:
        return [_criterio_dict_from_row(c) for c in programa.criterios_vinculados]
    return _criterios_from_json(programa)


def persist_programa_criterios(
    db: Session,
    programa: Programa,
    criterio_ids: list[int | str],
    limite_inferior: str = "",
    limite_superior: str = "",
    vigencia_inicio: str = "",
    vigencia_fim: str = "",
    prazo_indeterminado: bool = False,
) -> None:
    programa.criterios_vinculados.clear()
    db.flush()
    for raw in criterio_ids:
        if raw in (None, ""):
            continue
        catalogo = criterio_por_id(int(raw))
        if not catalogo:
            continue
        db.add(ProgramaCriterio(
            programa_id=programa.id,
            identificador_criterio=int(catalogo["id"]),
            nome=catalogo["nome"],
            categoria=catalogo.get("categoria"),
            tipo_dado=catalogo["tipo_dado"],
            limite_inferior=clean_text(limite_inferior) or None,
            limite_superior=clean_text(limite_superior) or None,
            vigencia_inicio=clean_text(vigencia_inicio) or None,
            vigencia_fim=None if prazo_indeterminado else (clean_text(vigencia_fim) or None),
            prazo_indeterminado=bool(prazo_indeterminado),
        ))


def migrate_programa_criterios_from_json(db: Session) -> None:
    programas = db.scalars(select(Programa)).all()
    for programa in programas:
        if programa.criterios_vinculados:
            continue
        for meta in _criterios_from_json(programa):
            cid = int(meta.get("identificadorCriterio") or meta.get("id") or 0)
            if cid <= 0:
                continue
            catalogo = criterio_por_id(cid) or {}
            db.add(ProgramaCriterio(
                programa_id=programa.id,
                identificador_criterio=cid,
                nome=meta.get("nome") or catalogo.get("nome") or f"Critério {cid}",
                categoria=meta.get("categoria") or catalogo.get("categoria"),
                tipo_dado=meta.get("tipoDado") or meta.get("tipo_dado") or catalogo.get("tipo_dado") or "Sim/Não",
                limite_inferior=meta.get("limiteInferior") or meta.get("limite_inferior"),
                limite_superior=meta.get("limiteSuperior") or meta.get("limite_superior"),
                vigencia_inicio=meta.get("vigenciaInicio") or meta.get("vigencia_inicio"),
                vigencia_fim=meta.get("vigenciaFim") or meta.get("vigencia_fim"),
                prazo_indeterminado=bool(meta.get("prazoIndeterminado") or meta.get("prazo_indeterminado")),
            ))


def folha_edicao_bloqueada(db: Session, folha: Folha) -> str | None:
    remessa = db.scalar(
        select(Remessa)
        .where(Remessa.folha_id == folha.id, Remessa.status.in_(["ENVIADA", "ACEITA"]))
        .order_by(Remessa.created_at.desc())
    )
    if remessa:
        return f"Folha bloqueada para edição: remessa marcada como {remessa.status}."
    return None


def validar_sequencial_nova_folha(
    db: Session,
    programa_id: int,
    ano: str,
    mes: int,
    tipo_folha: int,
    sequencial: int,
) -> None:
    if sequencial < 1:
        raise ValueError("O sequencial do arquivo deve ser no mínimo 1.")
    existentes = db.scalars(
        select(Folha.sequencial).where(
            Folha.programa_id == programa_id,
            Folha.ano == ano,
            Folha.mes == mes,
            Folha.tipo_folha == tipo_folha,
        )
    ).all()
    if sequencial == 1 and not existentes:
        return
    if sequencial > 1:
        anterior = db.scalar(
            select(Folha.id).where(
                Folha.programa_id == programa_id,
                Folha.ano == ano,
                Folha.mes == mes,
                Folha.tipo_folha == tipo_folha,
                Folha.sequencial == sequencial - 1,
            )
        )
        if not anterior:
            raise ValueError(
                f"Para usar sequencial {sequencial}, deve existir folha com sequencial {sequencial - 1} "
                f"na mesma competência, programa e tipo de folha."
            )


def criterios_from_program_columns(programa: Programa, form: dict[str, Any]) -> list[dict[str, Any]]:
    criterios = []
    for criterio in _programa_criterios(programa):
        cid = int(criterio.get("identificadorCriterio") or criterio.get("id") or 0)
        if cid <= 0:
            continue
        valor = clean_text(form.get(f"criterio_{cid}_valor"))
        aplicavel = clean_text(form.get(f"criterio_{cid}_aplicavel")).upper() or "S"
        if valor or aplicavel:
            criterios.append({
                "identificadorCriterio": cid,
                "valorAssociado": valor or "SIM",
                "aplicavel": "S" if aplicavel not in {"S", "N"} else aplicavel,
            })
    return criterios


def build_programa_criterios_json(
    criterio_ids: list[int | str],
    limite_inferior: str = "",
    limite_superior: str = "",
    vigencia_inicio: str = "",
    vigencia_fim: str = "",
    prazo_indeterminado: bool = False,
) -> str:
    criterios = []
    for raw in criterio_ids:
        if raw in (None, ""):
            continue
        catalogo = criterio_por_id(int(raw))
        if not catalogo:
            continue
        criterios.append({
            "identificadorCriterio": int(catalogo["id"]),
            "nome": catalogo["nome"],
            "categoria": catalogo["categoria"],
            "tipoDado": catalogo["tipo_dado"],
            "limiteInferior": clean_text(limite_inferior),
            "limiteSuperior": clean_text(limite_superior),
            "vigenciaInicio": clean_text(vigencia_inicio),
            "vigenciaFim": "" if prazo_indeterminado else clean_text(vigencia_fim),
            "prazoIndeterminado": bool(prazo_indeterminado),
        })
    return json.dumps(criterios, ensure_ascii=False)


def get_or_create_beneficiario(db: Session, data: dict[str, Any]) -> Beneficiario:
    cpf = digits_only(data.get("cpf"))
    beneficiario = db.scalar(select(Beneficiario).where(Beneficiario.cpf == cpf))
    if not beneficiario:
        beneficiario = Beneficiario(cpf=cpf)
        db.add(beneficiario)
    beneficiario.numero_nis = digits_only(data.get("numeroNIS")) or None
    beneficiario.nome = clean_text(data.get("nome")).upper()
    beneficiario.sexo = clean_text(data.get("sexo")).upper()
    beneficiario.data_nascimento = normalize_date(data.get("dataNascimento"))
    beneficiario.nacionalidade = clean_text(data.get("nacionalidade")) or "Brasil"
    beneficiario.nome_mae = clean_text(data.get("nomeMae")) or None
    beneficiario.endereco_cep = digits_only(data.get("enderecoCEP")) or None
    beneficiario.logradouro = clean_text(data.get("logradouro"))
    beneficiario.bairro = clean_text(data.get("bairro"))
    beneficiario.numero = clean_text(data.get("numero")) or None
    beneficiario.complemento = clean_text(data.get("complemento")) or None
    beneficiario.codigo_ibge_municipio = digits_only(data.get("codigoIBGEMunicipio"))
    return beneficiario


def _prepare_item_payload(form: dict[str, Any], programa: Programa | None = None) -> dict[str, Any]:
    errors = validar_beneficiario_basico(form)
    if errors:
        raise ValueError("; ".join(errors))
    criterios = parse_json_array(form.get("criterios_json"), "criterios", required=False)
    if not criterios and programa is not None:
        criterios = criterios_from_program_columns(programa, form)
    if not criterios:
        criterios = gerar_criterio_unico(form.get("criterio_id"), form.get("criterio_valor"), form.get("criterio_aplicavel"))
    c_errors = validar_criterios(criterios)
    if c_errors:
        raise ValueError("; ".join(c_errors))
    dependentes = parse_json_array(form.get("dependentes_json"), "dependentes", required=False)
    total_dependentes = int(form.get("totalDependentesBeneficio") or 0)
    d_errors = validar_dependentes(dependentes, total_dependentes)
    if d_errors:
        raise ValueError("; ".join(d_errors))
    valor = normalize_money(form.get("valorTotalTransferido"))
    total_pessoas = int(form.get("totalPessoasBeneficio") or 0)
    if total_pessoas < 1:
        raise ValueError("Total de pessoas no benefício deve ser no mínimo 1.")
    if total_pessoas != 1 + total_dependentes:
        raise ValueError("Total de pessoas deve ser 1 titular + quantidade de dependentes.")
    return {
        "criterios": criterios,
        "dependentes": dependentes,
        "valor": valor,
        "total_pessoas": total_pessoas,
        "total_dependentes": total_dependentes,
        "evidencia": clean_text(form.get("evidencia")) or None,
    }


def _find_existing_item(db: Session, folha: Folha, form: dict[str, Any]) -> FolhaItem | None:
    item_id = clean_text(form.get("item_id"))
    if item_id:
        item = db.get(FolhaItem, int(item_id))
        if item and item.folha_id == folha.id:
            return item
        raise ValueError(f"Registro com ID {item_id} não encontrado nesta folha.")
    cpf = digits_only(form.get("cpf"))
    if not cpf:
        return None
    return db.scalar(
        select(FolhaItem)
        .join(Beneficiario)
        .where(FolhaItem.folha_id == folha.id, Beneficiario.cpf == cpf)
    )


def make_item_from_form(db: Session, folha: Folha, form: dict[str, Any]) -> FolhaItem:
    payload = _prepare_item_payload(form, folha.programa)
    beneficiario = get_or_create_beneficiario(db, form)
    existing = _find_existing_item(db, folha, form)
    if existing:
        existing.beneficiario = beneficiario
        existing.valor_total_transferido = payload["valor"]
        existing.total_pessoas_beneficio = payload["total_pessoas"]
        existing.total_dependentes_beneficio = payload["total_dependentes"]
        existing.criterios_json = json.dumps(payload["criterios"], ensure_ascii=False)
        existing.dependentes_json = json.dumps(payload["dependentes"], ensure_ascii=False) if payload["dependentes"] else None
        existing.evidencia = payload["evidencia"]
        return existing
    item = FolhaItem(
        folha_id=folha.id,
        beneficiario=beneficiario,
        valor_total_transferido=payload["valor"],
        total_pessoas_beneficio=payload["total_pessoas"],
        total_dependentes_beneficio=payload["total_dependentes"],
        criterios_json=json.dumps(payload["criterios"], ensure_ascii=False),
        dependentes_json=json.dumps(payload["dependentes"], ensure_ascii=False) if payload["dependentes"] else None,
        evidencia=payload["evidencia"],
    )
    db.add(item)
    return item


def build_beneficio_json(item: FolhaItem) -> dict[str, Any]:
    b = item.beneficiario
    payload = {
        "cpf": b.cpf,
        "nome": b.nome,
        "sexo": b.sexo,
        "dataNascimento": b.data_nascimento,
        "nacionalidade": b.nacionalidade,
        "logradouro": b.logradouro,
        "bairro": b.bairro,
        "codigoIBGEMunicipio": b.codigo_ibge_municipio,
        "valorTotalTransferido": float(Decimal(item.valor_total_transferido).quantize(Decimal("0.01"))),
        "totalPessoasBeneficio": int(item.total_pessoas_beneficio),
        "totalDependentesBeneficio": int(item.total_dependentes_beneficio),
        "criterios": json.loads(item.criterios_json or "[]"),
    }
    optional = {
        "numeroNIS": b.numero_nis,
        "nomeMae": b.nome_mae,
        "enderecoCEP": b.endereco_cep,
        "numero": b.numero,
        "complemento": b.complemento,
    }
    for key, value in optional.items():
        if value:
            payload[key] = value
    dependentes = json.loads(item.dependentes_json or "[]")
    if dependentes:
        payload["dependentes"] = dependentes
    return payload


def build_folha_json(folha: Folha) -> dict[str, Any]:
    programa = folha.programa
    return {
        "unidadeAuditada": int(folha.unidade.codigo_etce),
        "anoReferencia": folha.ano,
        "mesReferencia": int(folha.mes),
        "tipoFolha": int(folha.tipo_folha),
        "sequencialArquivo": int(folha.sequencial),
        "programaSocial": int(programa.codigo_etce or 0),
        "beneficios": [build_beneficio_json(item) for item in folha.itens],
    }


def add_issue(db: Session, folha_id: int, item_id: int | None, severity: str, code: str, message: str) -> None:
    db.add(ValidationIssue(folha_id=folha_id, item_id=item_id, severity=severity, code=code, message=message))


def validate_folha(db: Session, folha: Folha) -> tuple[int, int]:
    db.execute(delete(ValidationIssue).where(ValidationIssue.folha_id == folha.id))
    db.flush()

    programa: Programa = folha.programa
    if not programa.codigo_etce or int(programa.codigo_etce) < 1:
        add_issue(db, folha.id, None, "BLOCK", "PROGRAMA_SEM_CODIGO_ETCE", "Cadastre o código do programa no e-TCERJ antes de gerar a folha.")
    if not programa.homologado_etce:
        add_issue(db, folha.id, None, "BLOCK", "PROGRAMA_NAO_HOMOLOGADO", "O programa precisa estar homologado no e-TCERJ.")
    if not programa.vigente:
        add_issue(db, folha.id, None, "BLOCK", "PROGRAMA_NAO_VIGENTE", "O programa não está marcado como vigente.")
    if not folha.itens:
        add_issue(db, folha.id, None, "BLOCK", "FOLHA_SEM_ITENS", "Inclua ao menos um beneficiário na folha.")

    if int(folha.ano) < 2026:
        add_issue(db, folha.id, None, "ALERT", "ANO_REFERENCIA_ANTIGO", "Ano de referência anterior a 2026: confira se a competência retroativa é permitida.")

    if folha.tipo_folha == 2:
        ordinaria = db.scalar(
            select(Folha).where(
                Folha.programa_id == folha.programa_id,
                Folha.ano == folha.ano,
                Folha.mes == folha.mes,
                Folha.tipo_folha == 1,
                Folha.id != folha.id,
            )
        )
        if not ordinaria:
            add_issue(
                db, folha.id, None, "BLOCK", "FOLHA_SUPLEMENTAR_SEM_ORDINARIA",
                "Folha suplementar (tipo 2) exige folha ordinária (tipo 1) na mesma competência e programa.",
            )

    if folha.sequencial > 1:
        anterior = db.scalar(
            select(Folha).where(
                Folha.programa_id == folha.programa_id,
                Folha.ano == folha.ano,
                Folha.mes == folha.mes,
                Folha.tipo_folha == folha.tipo_folha,
                Folha.sequencial == folha.sequencial - 1,
            )
        )
        if not anterior:
            add_issue(
                db, folha.id, None, "BLOCK", "SEQUENCIAL_INVALIDO",
                f"Sequencial {folha.sequencial} exige folha anterior com sequencial {folha.sequencial - 1}.",
            )

    remessa_mesma_comp = db.scalar(
        select(Remessa)
        .join(Folha)
        .where(
            Folha.programa_id == folha.programa_id,
            Folha.ano == folha.ano,
            Folha.mes == folha.mes,
            Folha.tipo_folha == folha.tipo_folha,
            Folha.sequencial == folha.sequencial,
            Folha.id != folha.id,
            Remessa.status.in_(["GERADA", "ENVIADA", "ACEITA"]),
        )
    )
    if remessa_mesma_comp:
        add_issue(
            db, folha.id, None, "BLOCK", "SEQUENCIAL_REUTILIZADO",
            "Já existe remessa para outra folha com o mesmo sequencial nesta competência. Use o próximo sequencial para correção.",
        )

    metadados_criterios = _programa_criterios(programa)

    cpf_count: dict[str, int] = {}
    for item in folha.itens:
        cpf_count[item.beneficiario.cpf] = cpf_count.get(item.beneficiario.cpf, 0) + 1
    for item in folha.itens:
        b = item.beneficiario
        data = {
            "cpf": b.cpf, "numeroNIS": b.numero_nis, "nome": b.nome, "sexo": b.sexo,
            "dataNascimento": b.data_nascimento, "nacionalidade": b.nacionalidade,
            "logradouro": b.logradouro, "bairro": b.bairro, "codigoIBGEMunicipio": b.codigo_ibge_municipio,
            "enderecoCEP": b.endereco_cep,
        }
        for msg in validar_beneficiario_basico(data):
            add_issue(db, folha.id, item.id, "BLOCK", "BENEFICIARIO_INVALIDO", msg)
        if cpf_count.get(b.cpf, 0) > 1:
            add_issue(db, folha.id, item.id, "BLOCK", "CPF_DUPLICADO_FOLHA", f"O CPF {b.cpf[:3]}*** já aparece mais de uma vez nesta folha.")
        if not cpf_valido(b.cpf):
            add_issue(db, folha.id, item.id, "BLOCK", "CPF_INVALIDO", "CPF com dígitos verificadores inválidos.")
        if not nacionalidade_deve_ser_pais(b.nacionalidade):
            add_issue(db, folha.id, item.id, "ALERT", "NACIONALIDADE_ADJETIVO", "Use o nome do país (ex.: Brasil), não o adjetivo (ex.: Brasileira).")
        try:
            criterios = json.loads(item.criterios_json or "[]")
        except Exception:
            criterios = []
            add_issue(db, folha.id, item.id, "BLOCK", "CRITERIO_INVALIDO", "Critérios em formato inválido.")
        for msg in validar_criterios(criterios):
            add_issue(db, folha.id, item.id, "BLOCK", "CRITERIO_INVALIDO", msg)
        for msg in validar_criterios_do_programa(criterios, metadados_criterios, folha.ano, int(folha.mes)):
            add_issue(db, folha.id, item.id, "BLOCK", "CRITERIO_PROGRAMA", msg)
        try:
            dependentes = json.loads(item.dependentes_json or "[]")
        except Exception:
            dependentes = []
            add_issue(db, folha.id, item.id, "BLOCK", "DEPENDENTE_INVALIDO", "Lista de dependentes em formato inválido.")
        for msg in validar_dependentes(dependentes, item.total_dependentes_beneficio):
            add_issue(db, folha.id, item.id, "BLOCK", "DEPENDENTE_INVALIDO", msg)
        if item.total_pessoas_beneficio != 1 + item.total_dependentes_beneficio:
            add_issue(db, folha.id, item.id, "BLOCK", "TOTAL_PESSOAS_INCOERENTE", "Total de pessoas deve ser 1 titular + dependentes informados.")
        if not item.evidencia:
            add_issue(db, folha.id, item.id, "ALERT", "SEM_EVIDENCIA", "Registre processo administrativo, despacho ou evidência do pagamento.")

    payload = build_folha_json(folha)
    validator = Draft7Validator(AUDFOBEN_SCHEMA, format_checker=FormatChecker())
    for error in sorted(validator.iter_errors(payload), key=lambda e: list(e.path)):
        path = ".".join(str(p) for p in error.path) or "arquivo"
        hint = friendly_schema_message(error.message)
        add_issue(db, folha.id, None, "BLOCK", "JSON_SCHEMA", f"Padrão AUDFOBEN — {path}: {hint}.")

    db.flush()
    block_count = db.scalar(select(func.count()).select_from(ValidationIssue).where(ValidationIssue.folha_id == folha.id, ValidationIssue.severity == "BLOCK")) or 0
    alert_count = db.scalar(select(func.count()).select_from(ValidationIssue).where(ValidationIssue.folha_id == folha.id, ValidationIssue.severity == "ALERT")) or 0
    folha.status = "VALIDADA" if block_count == 0 else "PENDENTE"
    db.commit()
    return int(block_count), int(alert_count)


def exportar_remessa(db: Session, folha: Folha, user: User | None) -> Remessa:
    block_count, _ = validate_folha(db, folha)
    if block_count:
        raise ValueError("A folha possui erros bloqueantes. Corrija as pendências antes de gerar o JSON.")
    payload = build_folha_json(folha)
    filename = f"AUDFOBEN_UG{folha.unidade.codigo_etce}_P{folha.programa.codigo_etce}_{folha.ano}_{folha.mes:02d}_T{folha.tipo_folha}_S{folha.sequencial}.json"
    path = EXPORT_DIR / filename
    raw = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
    path.write_bytes(raw)
    sha = hashlib.sha256(raw).hexdigest()
    remessa = Remessa(
        folha_id=folha.id,
        filename=filename,
        sha256=sha,
        size_bytes=len(raw),
        status="GERADA",
        created_by=user.id if user else None,
    )
    folha.status = "GERADA"
    db.add(remessa)
    db.commit()
    db.refresh(remessa)
    return remessa


def _row_is_empty(row: dict[str, Any]) -> bool:
    return not any(clean_text(row.get(col)) for col in REQUIRED_IMPORT_COLUMNS)


def _validate_headers(headers: list[str]) -> list[str]:
    missing = [col for col in REQUIRED_IMPORT_COLUMNS if col not in headers]
    if missing:
        return [f"Coluna obrigatória ausente na planilha: {', '.join(missing)}."]
    return []


def read_upload_table(filename: str, data: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    if len(data) > MAX_UPLOAD_SIZE:
        mb = MAX_UPLOAD_SIZE // (1024 * 1024)
        raise ValueError(f"Arquivo excede o limite de {mb} MB para importação.")
    lower = filename.lower()
    if not (lower.endswith(".csv") or lower.endswith(".xlsx")):
        raise ValueError("Formato inválido. Envie arquivo .csv ou .xlsx.")
    rows: list[dict[str, Any]] = []
    headers: list[str] = []
    if lower.endswith(".csv"):
        text = data.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        headers = [h.strip() for h in (reader.fieldnames or []) if h]
        header_errors = _validate_headers(headers)
        if header_errors:
            return header_errors, []
        for row in reader:
            rows.append({k.strip(): v for k, v in row.items() if k})
    else:
        if data[:2] != b"PK":
            raise ValueError("Arquivo Excel inválido ou corrompido.")
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        header_errors = _validate_headers(headers)
        if header_errors:
            return header_errors, []
        for values in ws.iter_rows(min_row=2, values_only=True):
            row = {headers[i]: values[i] for i in range(min(len(headers), len(values))) if headers[i]}
            rows.append(row)
    return [], rows


def importar_itens(db: Session, folha: Folha, filename: str, data: bytes) -> tuple[int, int, list[str]]:
    header_errors, rows = read_upload_table(filename, data)
    if header_errors:
        raise ValueError(" ".join(header_errors))
    imported = 0
    updated = 0
    errors: list[str] = []
    for idx, row in enumerate(rows, start=2):
        if _row_is_empty(row):
            continue
        try:
            form = dict(row)
            if not clean_text(form.get("criterios_json")):
                dinamicos = criterios_from_program_columns(folha.programa, form)
                if dinamicos:
                    form["criterios_json"] = json.dumps(dinamicos, ensure_ascii=False)
                else:
                    form["criterios_json"] = json.dumps(
                        gerar_criterio_unico(form.get("criterio_id"), form.get("criterio_valor"), form.get("criterio_aplicavel")),
                        ensure_ascii=False,
                    )
            had_item = bool(clean_text(form.get("item_id"))) or bool(
                db.scalar(
                    select(FolhaItem.id)
                    .join(Beneficiario)
                    .where(FolhaItem.folha_id == folha.id, Beneficiario.cpf == digits_only(form.get("cpf")))
                )
            )
            make_item_from_form(db, folha, form)
            if had_item:
                updated += 1
            else:
                imported += 1
        except Exception as exc:
            errors.append(f"Linha {idx}: {exc}")
    db.commit()
    return imported, updated, errors


def _exemplo_linha() -> dict[str, str]:
    return {
        "item_id": "",
        "cpf": "52998224725",
        "numeroNIS": "",
        "nome": "Beneficiario Exemplo",
        "sexo": "F",
        "dataNascimento": "1990-01-01",
        "nacionalidade": "Brasil",
        "nomeMae": "Mae Exemplo",
        "enderecoCEP": "28990000",
        "logradouro": "Rua Exemplo",
        "bairro": "Centro",
        "numero": "100",
        "complemento": "",
        "codigoIBGEMunicipio": "3305505",
        "valorTotalTransferido": "300.00",
        "totalPessoasBeneficio": "1",
        "totalDependentesBeneficio": "0",
        "criterio_id": "1",
        "criterio_valor": "Renda familiar dentro do critério do programa",
        "criterio_aplicavel": "S",
        "dependentes_json": "[]",
        "evidencia": "Processo administrativo nº 123/2026",
    }


def gerar_modelo_csv() -> bytes:
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=EXPECTED_COLUMNS)
    writer.writeheader()
    writer.writerow(_exemplo_linha())
    return output.getvalue().encode("utf-8-sig")


def gerar_modelo_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Beneficiarios"
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    header_font = Font(color="FFFFFF", bold=True)
    for col, name in enumerate(EXPECTED_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
    for col, name in enumerate(EXPECTED_COLUMNS, start=1):
        ws.cell(row=2, column=col, value=_exemplo_linha().get(name, ""))
    instr = wb.create_sheet("Instrucoes")
    instr["A1"] = "Como usar este modelo"
    instr["A3"] = "1. Preencha uma linha por beneficiário."
    instr["A4"] = "2. Para corrigir um registro existente, informe o item_id exportado da folha."
    instr["A5"] = "3. Se não houver item_id, o sistema atualiza pelo CPF quando já existir na folha."
    instr["A6"] = "4. dependentes_json deve ser uma lista JSON válida."
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def gerar_modelo_programa_xlsx(programa: Programa) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Beneficiarios"
    base_columns = [
        "item_id", "cpf", "numeroNIS", "nome", "sexo", "dataNascimento", "nacionalidade", "nomeMae", "enderecoCEP",
        "logradouro", "bairro", "numero", "complemento", "codigoIBGEMunicipio", "valorTotalTransferido",
        "totalPessoasBeneficio", "totalDependentesBeneficio", "dependentes_json", "evidencia",
    ]
    criterios = _programa_criterios(programa)
    dynamic_columns = []
    for criterio in criterios:
        cid = int(criterio.get("identificadorCriterio") or criterio.get("id") or 0)
        if cid:
            dynamic_columns.extend([f"criterio_{cid}_valor", f"criterio_{cid}_aplicavel"])
    columns = base_columns + dynamic_columns
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    header_font = Font(color="FFFFFF", bold=True)
    for col, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[cell.column_letter].width = max(14, min(35, len(name) + 2))
    exemplo = _exemplo_linha()
    for col, name in enumerate(base_columns, start=1):
        ws.cell(row=2, column=col, value=exemplo.get(name, ""))
    offset = len(base_columns) + 1
    for idx, criterio in enumerate(criterios):
        cid = int(criterio.get("identificadorCriterio") or criterio.get("id") or 0)
        ws.cell(row=2, column=offset + idx * 2, value="SIM")
        ws.cell(row=2, column=offset + idx * 2 + 1, value="S")

    instr = wb.create_sheet("Instrucoes")
    instr["A1"] = f"Modelo específico do programa: {programa.nome}"
    instr["A3"] = "Preencha uma linha por beneficiário. Não altere os nomes das colunas."
    instr["A4"] = "As colunas criterio_<ID>_valor e criterio_<ID>_aplicavel foram geradas a partir dos critérios vinculados ao programa."
    instr["A5"] = "Em aplicavel, use S ou N. Em valor, informe o dado que comprova o atendimento do critério, ex.: SIM, 200.00, 2, ou texto objetivo."
    instr["A7"] = "Critérios vinculados ao programa"
    headers = ["ID", "Nome", "Categoria", "Tipo de dado", "Limite inferior", "Limite superior", "Vigência", "Fim vigência"]
    for col, name in enumerate(headers, start=1):
        instr.cell(row=8, column=col, value=name).font = Font(bold=True)
    for row_idx, criterio in enumerate(criterios, start=9):
        instr.cell(row=row_idx, column=1, value=criterio.get("identificadorCriterio"))
        instr.cell(row=row_idx, column=2, value=criterio.get("nome"))
        instr.cell(row=row_idx, column=3, value=criterio.get("categoria"))
        instr.cell(row=row_idx, column=4, value=criterio.get("tipoDado") or criterio.get("tipo_dado"))
        instr.cell(row=row_idx, column=5, value=criterio.get("limiteInferior"))
        instr.cell(row=row_idx, column=6, value=criterio.get("limiteSuperior"))
        instr.cell(row=row_idx, column=7, value=criterio.get("vigenciaInicio"))
        instr.cell(row=row_idx, column=8, value="Prazo indeterminado" if criterio.get("prazoIndeterminado") else criterio.get("vigenciaFim"))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def exportar_folha_xlsx(folha: Folha) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"Folha_{folha.id}"
    for col, name in enumerate(EXPECTED_COLUMNS, start=1):
        ws.cell(row=1, column=col, value=name)
    for row_idx, item in enumerate(folha.itens, start=2):
        b = item.beneficiario
        criterios = json.loads(item.criterios_json or "[]")
        primeiro = criterios[0] if criterios else {}
        valores = {
            "item_id": str(item.id),
            "cpf": b.cpf,
            "numeroNIS": b.numero_nis or "",
            "nome": b.nome,
            "sexo": b.sexo,
            "dataNascimento": b.data_nascimento,
            "nacionalidade": b.nacionalidade,
            "nomeMae": b.nome_mae or "",
            "enderecoCEP": b.endereco_cep or "",
            "logradouro": b.logradouro,
            "bairro": b.bairro,
            "numero": b.numero or "",
            "complemento": b.complemento or "",
            "codigoIBGEMunicipio": b.codigo_ibge_municipio,
            "valorTotalTransferido": f"{Decimal(item.valor_total_transferido).quantize(Decimal('0.01'))}",
            "totalPessoasBeneficio": str(item.total_pessoas_beneficio),
            "totalDependentesBeneficio": str(item.total_dependentes_beneficio),
            "criterio_id": str(primeiro.get("identificadorCriterio", "")),
            "criterio_valor": str(primeiro.get("valorAssociado", "")),
            "criterio_aplicavel": str(primeiro.get("aplicavel", "")),
            "dependentes_json": item.dependentes_json or "[]",
            "evidencia": item.evidencia or "",
        }
        for col, name in enumerate(EXPECTED_COLUMNS, start=1):
            ws.cell(row=row_idx, column=col, value=valores.get(name, ""))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
